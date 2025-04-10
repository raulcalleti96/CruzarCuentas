import tkinter as tk
import webbrowser
from tkinter import filedialog, messagebox
from tkinter import ttk
import pandas as pd
import difflib
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import re
import subprocess

def normalizar(texto):
    if pd.isna(texto):
        return ""
    texto = str(texto).lower()
    texto = re.sub(r'[\W_]+', ' ', texto)
    texto = re.sub(r'\s+', ' ', texto).strip()
    return texto

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Cruce de Cuentas")
        self.root.geometry("1000x600")
        self.root.update_idletasks()
        width = self.root.winfo_screenwidth()
        height = self.root.winfo_screenheight()
        x = (width // 2) - (1000 // 2)
        y = (height // 2) - (600 // 2)
        self.root.geometry(f"1000x600+{x}+{y}")
        self.root.configure(highlightthickness=0, bd=0)

        self.clientes_path = ""
        self.facturas_path = ""

        self.estilos()
        self.init_ui()

    def estilos(self):
        style = ttk.Style()
        style.theme_use('default')
        style.configure("Rounded.TButton", font=("Segoe UI", 12, "bold"), padding=10, relief="flat", borderwidth=3)
        style.map("Rounded.TButton",
                  background=[('active', '#d1fae5'), ('!active', '#a7f3d0')],
                  foreground=[('pressed', '#111827'), ('active', '#111827')])
        style.configure("TLabel", font=("Segoe UI", 12))
        style.configure("Title.TLabel", font=("Segoe UI", 16, "bold"))
        style.configure("ResultBig.TLabel", font=("Segoe UI", 13, "bold"), foreground="white", background="#2563eb")

    def init_ui(self):
        self.left = tk.Frame(self.root, bg="#2563eb", width=500)
        self.left.pack(side="left", fill="both", expand=True)

        self.right = tk.Frame(self.root, bg="#fdfcf8", padx=30, pady=30, width=500)
        self.right.pack(side="right", fill="both", expand=True)
        self.bgcolor = self.right["bg"]

        # PANEL IZQUIERDO
        self.resultado1 = ttk.Label(self.left, text="", style="ResultBig.TLabel")
        self.resultado2 = ttk.Label(self.left, text="", style="ResultBig.TLabel")
        self.resultado3 = ttk.Label(self.left, text="", style="ResultBig.TLabel")
        self.resultado1.pack(pady=(100, 5))
        self.resultado2.pack()
        self.resultado3.pack()

        self.abrir_btn = ttk.Button(self.left, text="📂 Abrir archivo de facturas", command=self.abrir_archivo, style="Rounded.TButton")
        self.abrir_btn.pack(pady=20)
        self.abrir_btn.pack_forget()

        self.aviso_grande = tk.Label(self.left, text="⚠️ Solo se admiten archivos .xlsx",
                                     bg="#2563eb", fg="white", font=("Segoe UI", 14, "bold"))
        self.aviso_grande.pack(pady=30)
        # Línea del copyright
        licencia_label = tk.Label(
            self.left,
            text="Licencia MIT • © 2025",
            bg="#2563eb",
            fg="white",
            font=("Segoe UI", 9),
            cursor="hand2"
        )
        licencia_label.pack(side="bottom", pady=(0, 0))

        # Enlace a la licencia
        ver_licencia = tk.Label(
            self.left,
            text="Ver licencia",
            bg="#2563eb",
            fg="lightblue",
            font=("Segoe UI", 9, "underline"),
            cursor="hand2"
        )
        ver_licencia.pack(side="bottom")
        ver_licencia.bind("<Button-1>",
                          lambda e: webbrowser.open("https://github.com/raulcalleti96/CruzarCuentas/blob/main/LICENSE"))

        # PANEL DERECHO
        ttk.Label(self.right, text="📂 Introduce los archivos", style="Title.TLabel", background=self.bgcolor).pack(pady=(0, 20))

        ttk.Label(self.right, text="Clientes (.xlsx)", background=self.bgcolor).pack()
        ttk.Button(self.right, text="Seleccionar archivo de clientes", command=self.cargar_clientes, style="Rounded.TButton").pack(pady=5)
        self.check_clientes = ttk.Label(self.right, text="", background=self.bgcolor, foreground="green", font=("Segoe UI", 10, "bold"))
        self.check_clientes.pack()

        ttk.Label(self.right, text="Facturas (.xlsx)", background=self.bgcolor).pack(pady=(20, 0))
        ttk.Button(self.right, text="Seleccionar archivo de facturas", command=self.cargar_facturas, style="Rounded.TButton").pack(pady=5)
        self.check_facturas = ttk.Label(self.right, text="", background=self.bgcolor, foreground="green", font=("Segoe UI", 10, "bold"))
        self.check_facturas.pack()

        self.generar_btn = ttk.Button(self.right, text="⚙️ Generar", command=self.generar, style="Rounded.TButton")

    def mostrar_generar_si_listo(self):
        if self.clientes_path and self.facturas_path:
            self.generar_btn.pack(pady=20)

    def cargar_clientes(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if path.endswith(".xlsx"):
            self.clientes_path = path
            self.check_clientes.config(text="✅ Archivo de clientes cargado")
            self.mostrar_generar_si_listo()
        else:
            messagebox.showerror("Error", "El archivo debe ser .xlsx")

    def cargar_facturas(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if path.endswith(".xlsx"):
            self.facturas_path = path
            self.check_facturas.config(text="✅ Archivo de facturas cargado")
            self.mostrar_generar_si_listo()
        else:
            messagebox.showerror("Error", "El archivo debe ser .xlsx")

    def abrir_archivo(self):
        if os.name == 'posix':
            subprocess.call(('open', self.facturas_path))
        elif os.name == 'nt':
            os.startfile(self.facturas_path)

    def generar(self):
        if not self.clientes_path or not self.facturas_path:
            messagebox.showerror("Error", "Debes seleccionar ambos archivos.")
            return

        clientes = pd.read_excel(self.clientes_path, header=None)
        facturas = pd.read_excel(self.facturas_path, skiprows=6, header=None)

        clientes.columns = ['NumeroCuenta', 'Nombre']
        cuentas_dict = {
            normalizar(nombre): cuenta for nombre, cuenta in zip(clientes['Nombre'], clientes['NumeroCuenta'])
        }

        encontrados = 0
        dudosos = []
        no_encontrados = []

        wb = load_workbook(self.facturas_path)
        ws = wb.active

        rojo = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        naranja = PatternFill(start_color="FFFF9900", end_color="FFFF9900", fill_type="solid")
        negrita = Font(bold=True)

        for idx, nombre_raw in enumerate(facturas[4]):
            fila_excel = idx + 7
            nombre_normalizado = normalizar(nombre_raw)
            if nombre_normalizado in cuentas_dict:
                cuenta = cuentas_dict[nombre_normalizado]
                ws.cell(row=fila_excel, column=3).value = cuenta
                encontrados += 1
            else:
                posibles = difflib.get_close_matches(nombre_normalizado, cuentas_dict.keys(), n=1, cutoff=0.8)
                if posibles:
                    sugerido = posibles[0]
                    cuenta = cuentas_dict[sugerido]
                    ws.cell(row=fila_excel, column=3).value = cuenta
                    ws.cell(row=fila_excel, column=3).fill = naranja
                    ws.cell(row=fila_excel, column=3).font = negrita
                    dudosos.append((fila_excel, nombre_raw, sugerido))
                else:
                    ws.cell(row=fila_excel, column=3).fill = rojo
                    ws.cell(row=fila_excel, column=3).font = negrita
                    no_encontrados.append((fila_excel, nombre_raw))

        wb.save(self.facturas_path)

        self.resultado1.config(text=f"✅ {encontrados} cuentas encontradas.")
        self.resultado2.config(text=f"🟧 {len(dudosos)} dudosas en naranja.")
        self.resultado3.config(text=f"🟥 {len(no_encontrados)} sin coincidencia en rojo.")
        self.abrir_btn.pack()

def main():
    root = tk.Tk()
    App(root)
    root.mainloop()

if __name__ == "__main__":
    main()